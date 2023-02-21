import openpyxl
file="C:\selinium\Data\Empty.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]

for r in range(1,6):
    for c in range(1,3):
        sheet.cell(r,c).value="Data"


sheet.cell(3,3).value="NEW DATA"

workbook.save(file)


print("task completed sucessfully")
